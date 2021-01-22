# coding=utf-8

import openpyxl
import os

wb_in = openpyxl.Workbook()
wb_in.save('newdata.xlsx')
print("新建Excel文件成功")
filepath = "D:\\Zzl\\new_2018data"
dataset = os.listdir(filepath)
for i in dataset:
    print(i)
wb_in = openpyxl.load_workbook("newdata.xlsx")
sheets_in=wb_in.sheetnames
sheet_in=wb_in[sheets_in[0]]
k=1
count=1
for i in dataset:
    print("run the number: {0}      {1}".format(count,i))
    wb_read=openpyxl.load_workbook(filepath+"\\{}".format((i)))
    sheets_read=wb_read.sheetnames
    sheet_read=wb_read[sheets_read[0]]
    title=sheet_read.cell(row=1,column=1).value
    row_item=2
    target=0
    for l in range(1,8):
        block=sheet_read.cell(row=l,column=1).value

        if isinstance(block,str):
            if block.find("工业销售产值")!= -1:
                target=l
                break
     #if title.find("主要经济效益指标完成情况")!= -1 & target!=0:
    if  target!=0:

        for j in range(1,7) :
            print(sheet_read.cell(row=row_item,column=j).value,end=" ")
        print("\n")
        head=title.find("年")
        tail=title.find("电子")

        sheet_in.cell(row=count,column=1).value=title[head+1:tail-1]
        sheet_in.cell(row=count, column=2).value =sheet_read.cell(row=l,column=3).value
        sheet_in.cell(row=count, column=3).value =sheet_read.cell(row=l,column=4).value
        sheet_in.cell(row=count, column=4).value = sheet_read.cell(row=l, column=5).value
        count = count + 1

    wb_read.close()
wb_in.save('newdata.xlsx')
wb_in.close()
print("Finished")
